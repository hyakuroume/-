#! python3
# updateProduce.py - 農産物スプレッドシートの価格を訂正する

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

#農産物の種類と、更新する価格
PRICE_UPDATES = {'Garlic':3.07,
                 'Celery':1.19,
                 'Le,on':1.27}

for row_num in range(2, sheet.max_row + 1): #先頭行をスキップ
    produce_name = sheet.cell(row = row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row = row_num, column=2).value = PRICE_UPDATES[product_name]

wb.save('updatedProduceSales.xlsx')

import openpyxl
import datetime
import calendar
from datetime import date

name_list = ["谷佳祐","谷口浩一","谷口洋満","德永弘美","徳本直也","内藤登","長岡克幸","中村遥香","野原郁映"]
weekday_list = ["月","火","水","木","金","土","日"]

d = datetime.datetime.today()
year = d.year
month = d.month

#月末の長さを取得
lastday_len = calendar.monthrange(year,month)[1]

#Ｅｘｃｅｌを起動
wb = openpyxl.Workbook()
ws =wb.active

#月を入力
ws["A1"].value = str(month) + "月"

#日数を入力
for i in range(lastday_len):
    ws.cell(i+2,1,value=i+1)

#曜日を入力
for i in range(lastday_len):
    #月初めの曜日を求める
    date1 = date(year,month,i+1)
    dw = date1.weekday()
    
    ws.cell(i+2,2,value=weekday_list[dw])


#名前を入力
for i in range(len(name_list)):
    ws.cell(1,i+3,value=name_list[i])

#Ｅｘｃｅｌを保存
wb.save("test.xlsx")
    
